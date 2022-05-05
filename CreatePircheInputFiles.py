import argparse
from openpyxl import load_workbook, Workbook
from os import makedirs
from os.path import join, isdir


def readDsaFile(dsaFile=None):
    print('Reading dsaFile:' + str(dsaFile))

    # Read the excel
    xlWorkbook = load_workbook(filename=dsaFile)

    typingTab = xlWorkbook['Typing']
    positiveNegativeTab = xlWorkbook['3000']

    genotypes={}
    dsaIndicators={}

    for rowIndexRaw, row in enumerate(typingTab.iter_rows()):
        if rowIndexRaw > 0:
            transplantationId = str(row[0].value).strip()
            currentGenotypes = {}
            currentDsaIndicators = {}


            currentGenotypes['rA1'] = str(row[1].value).strip()
            currentGenotypes['rA2'] = str(row[2].value).strip()
            currentGenotypes['rB1'] = str(row[3].value).strip()
            currentGenotypes['rB2'] = str(row[4].value).strip()
            currentGenotypes['rC1'] = str(row[5].value).strip()
            currentGenotypes['rC2'] = str(row[6].value).strip()
            currentGenotypes['rDRB11'] = str(row[7].value).strip()
            currentGenotypes['rDRB12'] = str(row[8].value).strip()
            currentGenotypes['rDRB31'] = str(row[9].value).strip()
            currentGenotypes['rDRB32'] = str(row[10].value).strip()
            currentGenotypes['rDRB41'] = str(row[11].value).strip()
            currentGenotypes['rDRB42'] = str(row[12].value).strip()
            currentGenotypes['rDRB51'] = str(row[13].value).strip()
            currentGenotypes['rDRB52'] = str(row[14].value).strip()
            currentGenotypes['rDQB11'] = str(row[15].value).strip()
            currentGenotypes['rDQB12'] = str(row[16].value).strip()
            currentGenotypes['rDQA11'] = str(row[17].value).strip()
            currentGenotypes['rDQA12'] = str(row[18].value).strip()
            currentGenotypes['rDPB11'] = str(row[19].value).strip()
            currentGenotypes['rDPB12'] = str(row[20].value).strip()
            currentGenotypes['rDPA11'] = str(row[21].value).strip()
            currentGenotypes['rDPA12'] = str(row[22].value).strip()
            currentGenotypes['rDRB3451'] = str(row[23].value).strip()
            currentGenotypes['rDRB3452'] = str(row[24].value).strip()

            currentGenotypes['dA1'] = str(row[33].value).strip()
            currentGenotypes['dA2'] = str(row[34].value).strip()
            currentGenotypes['dB1'] = str(row[35].value).strip()
            currentGenotypes['dB2'] = str(row[36].value).strip()
            currentGenotypes['dC1'] = str(row[37].value).strip()
            currentGenotypes['dC2'] = str(row[38].value).strip()
            currentGenotypes['dDRB11'] = str(row[39].value).strip()
            currentGenotypes['dDRB12'] = str(row[40].value).strip()
            currentGenotypes['dDRB31'] = str(row[41].value).strip()
            currentGenotypes['dDRB32'] = str(row[42].value).strip()
            currentGenotypes['dDRB41'] = str(row[43].value).strip()
            currentGenotypes['dDRB42'] = str(row[44].value).strip()
            currentGenotypes['dDRB51'] = str(row[45].value).strip()
            currentGenotypes['dDRB52'] = str(row[46].value).strip()
            currentGenotypes['dDQB11'] = str(row[47].value).strip()
            currentGenotypes['dDQB12'] = str(row[48].value).strip()
            currentGenotypes['dDQA11'] = str(row[49].value).strip()
            currentGenotypes['dDQA12'] = str(row[50].value).strip()
            currentGenotypes['dDPB11'] = str(row[51].value).strip()
            currentGenotypes['dDPB12'] = str(row[52].value).strip()
            currentGenotypes['dDPA11'] = str(row[53].value).strip()
            currentGenotypes['dDPA12'] = str(row[54].value).strip()
            currentGenotypes['dDRB3451'] = str(row[55].value).strip()
            currentGenotypes['dDRB3452'] = str(row[56].value).strip()

            # The "cell" indexing here is 1-based
            currentDsaIndicators['a1'] = positiveNegativeTab.cell(rowIndexRaw+1, 31).value
            currentDsaIndicators['a2'] = positiveNegativeTab.cell(rowIndexRaw + 1, 32).value
            currentDsaIndicators['b1'] = positiveNegativeTab.cell(rowIndexRaw+1, 33).value
            currentDsaIndicators['b2'] = positiveNegativeTab.cell(rowIndexRaw + 1, 34).value
            currentDsaIndicators['c1'] = positiveNegativeTab.cell(rowIndexRaw+1, 35).value
            currentDsaIndicators['c2'] = positiveNegativeTab.cell(rowIndexRaw + 1, 36).value
            currentDsaIndicators['drb11'] = positiveNegativeTab.cell(rowIndexRaw+1, 37).value
            currentDsaIndicators['drb12'] = positiveNegativeTab.cell(rowIndexRaw + 1, 38).value
            currentDsaIndicators['drb3451'] = positiveNegativeTab.cell(rowIndexRaw+1, 39).value
            currentDsaIndicators['drb3452'] = positiveNegativeTab.cell(rowIndexRaw + 1, 40).value
            currentDsaIndicators['dq1'] = positiveNegativeTab.cell(rowIndexRaw+1, 41).value
            currentDsaIndicators['dq2'] = positiveNegativeTab.cell(rowIndexRaw + 1, 42).value
            currentDsaIndicators['dq3'] = positiveNegativeTab.cell(rowIndexRaw+1, 43).value
            currentDsaIndicators['dq4'] = positiveNegativeTab.cell(rowIndexRaw + 1, 44).value
            currentDsaIndicators['dp1'] = positiveNegativeTab.cell(rowIndexRaw+1, 45).value
            currentDsaIndicators['dp2'] = positiveNegativeTab.cell(rowIndexRaw + 1, 46).value
            currentDsaIndicators['dp3'] = positiveNegativeTab.cell(rowIndexRaw+1, 47).value
            currentDsaIndicators['dp4'] = positiveNegativeTab.cell(rowIndexRaw + 1, 48).value

            currentDsaIndicators['predsa'] = positiveNegativeTab.cell(rowIndexRaw + 1, 18).value
            currentDsaIndicators['dndsa'] = positiveNegativeTab.cell(rowIndexRaw + 1, 59).value
            currentDsaIndicators['post_self'] = positiveNegativeTab.cell(rowIndexRaw + 1, 28).value
            currentDsaIndicators['pre_self'] = positiveNegativeTab.cell(rowIndexRaw + 1, 8).value

            '''
            IF preDSA_3000 & dnDSA_3000 = 1 or post_self_3000 = 1 then exclude
            I ignored the preTx_self_3000 with high values for now
            No I only excluded cases where there was preDSA and postTx DSA because then you are not sure that you look at immunisation of transplantation
            '''
            if(str(currentDsaIndicators['predsa']).strip()=="1" and str(currentDsaIndicators['dndsa']).strip()=="1"):
                print('transplantationId ' + str(transplantationId) + ' rejected because pre-dsa and denovo dsa are both positive')
            elif(str(currentDsaIndicators['post_self']).strip()=="1"):
                print('transplantationId ' + str(transplantationId) + ' rejected because post-self antibodies are positive')
            elif(len(str(currentGenotypes['rA1']).strip())<3):
                print('transplantationId ' + str(transplantationId) + ' rejected because missing HLA-A genotype')
            #elif (str(currentDsaIndicators['pre_self']).strip() == "1"):
            #    print('transplantationId ' + str(transplantationId) + ' rejected because pre-self antibodies are positive')
            else:
                genotypes[transplantationId] = currentGenotypes
                dsaIndicators[transplantationId] = currentDsaIndicators

    return genotypes, dsaIndicators


def addTyping(typingLine=None, newGenotype=None, delimiter=','):
    if(newGenotype not in typingLine
        and '?' not in newGenotype
        and len(newGenotype.strip())>1):
        typingLine= typingLine+delimiter+newGenotype.strip()
    return typingLine


def addTypingConditional(newGenotype=None, donorLine=None, donorPositiveLine=None, donorNegativeLine=None, dsaConditional=None):
    #Check if this bead was positive, leading to a positive response.
    donorLine = addTyping(typingLine=donorLine, newGenotype=newGenotype)
    if(str(dsaConditional) == '1'):
        donorPositiveLine = addTyping(typingLine=donorPositiveLine, newGenotype=newGenotype)
    elif(str(dsaConditional) == '0'):
        donorNegativeLine = addTyping(typingLine=donorNegativeLine, newGenotype=newGenotype)
    else:
        pass# This means its probably '9999' and we are not sure what to do with it.
    return donorLine, donorPositiveLine, donorNegativeLine


def createPirchePairLine(transplantationId=None, genotype=None, dsaIndicator=None, newline='\n'):
    patientLine = 'Recipient_' + str(transplantationId)
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rA1'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rA2'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rB1'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rB2'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rC1'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rC2'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB11'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB12'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB31'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB32'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB41'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB42'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB51'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB52'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDQB11'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDQB12'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDQA11'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDQA12'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDPB11'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDPB12'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDPA11'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDPA12'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB3451'])
    patientLine = addTyping(typingLine=patientLine, newGenotype=genotype['rDRB3452'])

    donorLine = 'Donor_' + str(transplantationId) + '_All'
    donorPositiveLine = 'Donor_' + str(transplantationId) + '_Positive'
    donorNegativeLine = 'Donor_' + str(transplantationId) + '_Negative'

    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dA1'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['a1'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dA2'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['a2'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dB1'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['b1'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dB2'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['b2'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dC1'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['c1'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dC2'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['c2'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDRB11'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['drb11'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDRB12'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['drb12'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDRB3451'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['drb3451'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDRB3452'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['drb3452'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDQB11'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['dq1'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDQB12'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['dq2'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDQA11'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['dq3'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDQA12'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['dq4'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDPB11'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['dp1'])

    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDPB12'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['dp2'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDPA11'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['dp3'])
    donorLine, donorPositiveLine, donorNegativeLine = addTypingConditional(newGenotype=genotype['dDPA12'], donorLine=donorLine
        , donorPositiveLine=donorPositiveLine, donorNegativeLine=donorNegativeLine, dsaConditional=dsaIndicator['dp4'])

    positiveAlleles = str(donorPositiveLine)
    # The donor lines need to have at a minimum A,B,DRB1 types.
    # If the're missing, fill it in using the recipient typing (should not make any new pirches I believe)
    if('A*' not in donorPositiveLine):
        donorPositiveLine = addTyping(typingLine=donorPositiveLine, newGenotype=genotype['rA1'])
    if('B*' not in donorPositiveLine):
        donorPositiveLine = addTyping(typingLine=donorPositiveLine, newGenotype=genotype['rB1'])
    if('DRB1*' not in donorPositiveLine):
        donorPositiveLine = addTyping(typingLine=donorPositiveLine, newGenotype=genotype['rDRB11'])

    if ('A*' not in donorNegativeLine):
        donorNegativeLine = addTyping(typingLine=donorNegativeLine, newGenotype=genotype['rA1'])
    if ('B*' not in donorNegativeLine):
        donorNegativeLine = addTyping(typingLine=donorNegativeLine, newGenotype=genotype['rB1'])
    if ('DRB1*' not in donorNegativeLine):
        donorNegativeLine = addTyping(typingLine=donorNegativeLine, newGenotype=genotype['rDRB11'])

    pirchPairLine=patientLine + newline + donorLine + newline + donorPositiveLine + newline + donorNegativeLine
    return pirchPairLine, positiveAlleles


def createPircheInputFiles(genotypes=None, dsaIndicators=None, outputDir=None, batchSize=100, newline='\n', delimiter=','):
    pircheDir = join(outputDir, 'PIRCHE_InputFiles')
    print('Creating PIRCHE Input Files:' + str(pircheDir))
    if not isdir(pircheDir):
        makedirs(pircheDir)

    positiveAllelesOutputFileName = join(outputDir, 'PositiveAlleles.csv')
    positiveAllelesOutputFile = open(positiveAllelesOutputFileName, 'w')

    batchIndex = 1
    currentBatchCount = 1
    outputFileName = join(pircheDir, 'PircheInputBatch_' + str(batchIndex) + '.csv')
    outputFile = open(outputFileName, 'w')

    for index, transplantationId in enumerate(list(genotypes.keys())):
        pirchePairLine, positiveAlleles = createPirchePairLine(transplantationId=transplantationId, genotype=genotypes[transplantationId], dsaIndicator=dsaIndicators[transplantationId])
        outputFile.write(pirchePairLine+newline)
        positiveAllelesOutputFile.write(positiveAlleles + newline)

        currentBatchCount += 1

        if(currentBatchCount>int(batchSize)):
            outputFile.close()
            batchIndex += 1
            currentBatchCount = 1
            outputFileName = join(pircheDir, 'PircheInputBatch_' + str(batchIndex) + '.csv')
            outputFile = open(outputFileName, 'w')
        else:
            outputFile.write(delimiter + newline)

    outputFile.close()
    positiveAllelesOutputFile.close()

if __name__ == '__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument("-v", "--verbose", help="verbose operation", action="store_true")
    parser.add_argument("-d", "--dsa", help="summary excel file for dsa", required=True)
    parser.add_argument("-o", "--output", help="output dir", required=True)

    args = parser.parse_args()
    verbose = args.verbose

    if not isdir(args.output):
        makedirs(args.output)

    genotypes, dsaIndicators = readDsaFile(dsaFile=args.dsa)

    createPircheInputFiles(genotypes=genotypes, dsaIndicators=dsaIndicators, outputDir=args.output)




